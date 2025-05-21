import openpyxl
import csv
from tkinter import filedialog as filedialog
from tkinter.messagebox import showwarning, showinfo
import json


class virketall:
    showinfo("Virketall","Husk å legge kopiere verdiene slik at formelene ikke er med i excel filen\nog husk å split merge headinger")
    def __init__(self):
        self.wb = None
        self.ws = None
        self.config = None
        self.choosenFile = None

        self.getConfigFile()
        if self.config:
            self.getFileByname()
        if self.choosenFile:
            self.readExcel()
            self.tablesToRead = self.config.get("tablesToRead") or None
            self.startRow = self.config.get("startrow") or None
            self.columns = self.config.get("columns") or None
            self.numberOfCategories = self.config.get("numberOfCategories") or None
            self.categoryNameColumn = self.config.get("categoryNameColumn") or None
            self.headingsOffsetFromStartRow = (
                self.config.get("headingsOffsetFromStartRow") or None
            )
        if self.wb and self.ws:
            self.getValues()
            self.writeToCSV()

    def getConfigFile(self):
        try:
            with open("./config.json", encoding="utf8") as f:
                try:
                    data = json.load(f)
                    self.config = data
                except ValueError:
                    showwarning(
                        title="Parsing error",
                        message="Could not parse the JSON config, it needs to be fixed",
                    )
                return
        except FileNotFoundError:
            showwarning(
                "Could not find config file",
                message="Config file not found please ensure its present",
            )

    def checkIfFileIsChoosen(self):
        if self.choosenFile:
            pass
        else:
            self.getFileByname()

    def getFileByname(self):
        self.choosenFile = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=(("Excel", "*.xlsx"),),
            title="Choose file with new virketall",
        )

    def readExcel(self):
        self.wb = openpyxl.load_workbook(self.choosenFile) or None
        self.ws = self.wb[self.config.get("sheetName")] or None

    def getValues(self):
        self.valuesList = []
        for table in self.tablesToRead:
            for column in self.columns:
                headers = [
                    self.ws[f"{column}{table.get("startRow") + headerNr}"].value
                    for headerNr in self.headingsOffsetFromStartRow
                ]
                for i in range(self.numberOfCategories):
                    currentRow = table.get("startRow") + i
                    currentValue = self.ws[f"{column}{currentRow}"].value
                    currentCategory = self.ws[
                        f"{self.categoryNameColumn}{currentRow}"
                    ].value
                    self.addValuesToList(
                        headers=headers,
                        currentValue=currentValue,
                        currentCategory=currentCategory,
                        actor=table.get("tableName"),
                    )

    def addValuesToList(self, actor, headers, currentValue, currentCategory):
        tempLst = []
        [tempLst.append(header) for header in headers]
        tempLst.append(actor)
        tempLst.append(currentCategory)
        tempLst.append(currentValue)
        self.valuesList.append(tempLst)

    def writeToCSV(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",)
        if not filename:
            showwarning(
                title="No file choosen",
                message="Please choose a file to save the data to",
            )
            return
        with open(filename, "w", newline="",encoding="UTF8") as f:
            writer = csv.writer(f)
            writer.writerow(self.config.get("headings"))
            for row in self.valuesList:
                writer.writerow(row)
        print("CSV file created successfully!")


if __name__ == "__main__":
    virketall()
